import openpyxl, os

print(os.getcwd())

# open excell file | called a workbook in excel
workbook = openpyxl.load_workbook('excel-example.xlsx')

# open tab | called a sheet in excel
sheet = workbook.get_sheet_by_name('Sheet1')

# return sheet names
# print(workbook.get_sheet_names())

# return cell object in cell A1
cell = sheet['A1']

# return value from cell object
print(cell.value)

# change data type
print(str(cell.value))

# call cell by number
sheet.cell(row=1, column=2)

# method can be used in for loop
for i in range(1,8):
  print(i, sheet.cell(row=i, column=2).value)

# creates new workbook
wb = openpyxl.Workbook()

# search for sheet and create sheet variable for it
# print(wb.get_sheet_by_name())
sheet = wb.get_sheet_by_name('Sheet')

# input some values
sheet['A1'].value == 42
sheet['A2'].value == 'Hello'

# add more sheets one on index 0
sheet2 = wb.create_sheet()
sheet3 = wb.create_sheet(index=0, title='My other new sheet')

# get title new sheet
print(sheet2.title)

# change title of sheet
sheet2.title = 'My new sheet name'

# save (document_name)
wb.save('example.xlsx')
